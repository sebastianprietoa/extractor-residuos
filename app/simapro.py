from __future__ import annotations

import ast
import csv
import io
import re
import unicodedata
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


CSV_DELIMITER = ";"

SECTION_LABELS = {
    "Products": "Producto principal",
    "Avoided products": "Productos evitados",
    "Resources": "Recursos",
    "Materials/fuels": "Materiales/combustibles",
    "Electricity/heat": "Electricidad/calor",
    "Emissions to air": "Emisiones al aire",
    "Emissions to water": "Emisiones al agua",
    "Emissions to soil": "Emisiones al suelo",
    "Final waste flows": "Flujos finales de residuos",
    "Non material emissions": "Emisiones no materiales",
    "Social issues": "Aspectos sociales",
    "Economic issues": "Aspectos economicos",
    "Waste to treatment": "Residuos a tratamiento",
    "Waste treatment": "Tratamiento de residuos",
    "Input parameters": "Parametros de entrada",
    "Calculated parameters": "Parametros calculados",
}


def _normalize_label(value: Any) -> str:
    if value is None:
        text = ""
    else:
        text = str(value).strip().replace("\n", " ")
        text = re.sub(r"\s+", " ", text)
    text = text.casefold()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text.strip()


SECTION_LABEL_ALIASES = {
    _normalize_label("Products"): "Products",
    _normalize_label("Productos"): "Products",
    _normalize_label("Avoided products"): "Avoided products",
    _normalize_label("Productos evitados"): "Avoided products",
    _normalize_label("Resources"): "Resources",
    _normalize_label("Recursos"): "Resources",
    _normalize_label("Materials/fuels"): "Materials/fuels",
    _normalize_label("Materials and fuels"): "Materials/fuels",
    _normalize_label("Materiales/combustibles"): "Materials/fuels",
    _normalize_label("Electricity/heat"): "Electricity/heat",
    _normalize_label("Electricidad/calor"): "Electricity/heat",
    _normalize_label("Emissions to air"): "Emissions to air",
    _normalize_label("Emisiones al aire"): "Emissions to air",
    _normalize_label("Emissions to water"): "Emissions to water",
    _normalize_label("Emisiones al agua"): "Emissions to water",
    _normalize_label("Emissions to soil"): "Emissions to soil",
    _normalize_label("Emisiones al suelo"): "Emissions to soil",
    _normalize_label("Final waste flows"): "Final waste flows",
    _normalize_label("Flujos finales de residuos"): "Final waste flows",
    _normalize_label("Non material emissions"): "Non material emissions",
    _normalize_label("Emisiones no materiales"): "Non material emissions",
    _normalize_label("Social issues"): "Social issues",
    _normalize_label("Asuntos sociales"): "Social issues",
    _normalize_label("Economic issues"): "Economic issues",
    _normalize_label("Asuntos economicos"): "Economic issues",
    _normalize_label("Waste to treatment"): "Waste to treatment",
    _normalize_label("Residuos a tratamiento"): "Waste to treatment",
    _normalize_label("Waste treatment"): "Waste treatment",
    _normalize_label("Tratamiento de residuos"): "Waste treatment",
    _normalize_label("Input parameters"): "Input parameters",
    _normalize_label("Parametros de entrada"): "Input parameters",
    _normalize_label("Calculated parameters"): "Calculated parameters",
    _normalize_label("Parametros calculados"): "Calculated parameters",
}

METADATA_FIELD_ALIASES = {
    _normalize_label("Category type"): "Category type",
    _normalize_label("Tipo de categoria"): "Category type",
    _normalize_label("Process identifier"): "Process identifier",
    _normalize_label("Identificador del proceso"): "Process identifier",
    _normalize_label("Type"): "Type",
    _normalize_label("Tipo"): "Type",
    _normalize_label("Process name"): "Process name",
    _normalize_label("Nombre del proceso"): "Process name",
    _normalize_label("Status"): "Status",
    _normalize_label("Estado"): "Status",
    _normalize_label("Time period"): "Time period",
    _normalize_label("Periodo"): "Time period",
    _normalize_label("Geography"): "Geography",
    _normalize_label("Geografia"): "Geography",
    _normalize_label("Technology"): "Technology",
    _normalize_label("Tecnologia"): "Technology",
    _normalize_label("Representativeness"): "Representativeness",
    _normalize_label("Representatividad"): "Representativeness",
    _normalize_label("Multiple output allocation"): "Multiple output allocation",
    _normalize_label("Substitution allocation"): "Substitution allocation",
    _normalize_label("Cut off rules"): "Cut off rules",
    _normalize_label("Capital goods"): "Capital goods",
    _normalize_label("Boundary with nature"): "Boundary with nature",
    _normalize_label("Infrastructure"): "Infrastructure",
    _normalize_label("Date"): "Date",
    _normalize_label("Fecha"): "Date",
    _normalize_label("Record"): "Record",
    _normalize_label("Registro"): "Record",
    _normalize_label("Generator"): "Generator",
    _normalize_label("Generador"): "Generator",
    _normalize_label("External documents"): "External documents",
    _normalize_label("Literature references"): "Literature references",
    _normalize_label("Collection method"): "Collection method",
    _normalize_label("Data treatment"): "Data treatment",
    _normalize_label("Tratamiento de datos"): "Data treatment",
    _normalize_label("Verification"): "Verification",
    _normalize_label("Verificacion"): "Verification",
    _normalize_label("Comment"): "Comment",
    _normalize_label("Comentario"): "Comment",
    _normalize_label("Allocation rules"): "Allocation rules",
    _normalize_label("System description"): "System description",
    _normalize_label("Descripcion del sistema"): "System description",
}

OUTPUT_COLUMNS = [
    "carpeta",
    "subcarpeta",
    "archivo",
    "ruta_relativa",
    "process_identifier",
    "fecha_proceso",
    "producto_principal",
    "cantidad_producto",
    "unidad_producto",
    "dataset_utilizado",
    "clasificacion",
    "seccion_origen",
    "compartimento",
    "cantidad_original",
    "cantidad_normalizada",
    "unidad_medida",
    "incertidumbre",
    "comentario",
    "fila_origen",
]

EXCEL_COLUMNS = [
    ("Carpeta", "carpeta"),
    ("Producto principal", "producto_principal"),
    ("Dataset utilizado", "dataset_utilizado"),
    ("Clasificacion", "clasificacion"),
    ("Seccion origen", "seccion_origen"),
    ("Compartimiento", "compartimento"),
    ("Cantidad normalizada", "cantidad_normalizada"),
    ("Unidad de medida", "unidad_medida"),
]


@dataclass(frozen=True)
class UploadedWorkbook:
    filename: str
    content: bytes


@dataclass(frozen=True)
class ExtractionResult:
    rows: list[dict[str, Any]]
    errors: list[dict[str, str]]
    summary: dict[str, int]


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().replace("\n", " ")
    return re.sub(r"\s+", " ", text)


def normalize_decimal_text(value: str) -> str:
    return re.sub(r"(?<=\d),(?=\d)", ".", value)


class _SafeExpressionEvaluator(ast.NodeVisitor):
    allowed_binops = {
        ast.Add: lambda a, b: a + b,
        ast.Sub: lambda a, b: a - b,
        ast.Mult: lambda a, b: a * b,
        ast.Div: lambda a, b: a / b,
    }
    allowed_unary = {
        ast.UAdd: lambda a: a,
        ast.USub: lambda a: -a,
    }

    def visit_Expression(self, node: ast.Expression) -> float:
        return self.visit(node.body)

    def visit_Constant(self, node: ast.Constant) -> float:
        if isinstance(node.value, (int, float)):
            return float(node.value)
        raise ValueError("Constante no numerica.")

    def visit_BinOp(self, node: ast.BinOp) -> float:
        operation = self.allowed_binops.get(type(node.op))
        if operation is None:
            raise ValueError("Operacion no permitida.")
        return operation(self.visit(node.left), self.visit(node.right))

    def visit_UnaryOp(self, node: ast.UnaryOp) -> float:
        operation = self.allowed_unary.get(type(node.op))
        if operation is None:
            raise ValueError("Operacion unaria no permitida.")
        return operation(self.visit(node.operand))

    def generic_visit(self, node: ast.AST) -> float:
        raise ValueError(f"Expresion no permitida: {type(node).__name__}")


def parse_quantity(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = normalize_decimal_text(str(value).strip())
    if not text:
        return None
    if not re.fullmatch(r"[0-9eE+\-*/().\s]+", text):
        return None

    try:
        tree = ast.parse(text, mode="eval")
        result = _SafeExpressionEvaluator().visit(tree)
    except Exception:
        return None

    if result == float("inf") or result == float("-inf"):
        return None
    return result


def format_quantity(value: float | None) -> str:
    if value is None:
        return ""
    if value.is_integer():
        return str(int(value))
    return f"{value:.12g}"


def split_uploaded_path(filename: str) -> dict[str, str]:
    normalized = filename.replace("\\", "/").strip("/")
    parts = [part for part in normalized.split("/") if part]
    if parts and parts[0].lower() == "recursos":
        parts = parts[1:]

    archivo = parts[-1] if parts else Path(filename).name
    folder_parts = parts[:-1]
    carpeta = folder_parts[0] if folder_parts else ""
    subcarpeta = " / ".join(folder_parts[1:]) if len(folder_parts) > 1 else ""
    ruta_relativa = "/".join(parts) if parts else archivo

    return {
        "carpeta": carpeta,
        "subcarpeta": subcarpeta,
        "archivo": archivo,
        "ruta_relativa": ruta_relativa,
    }


def _row_has_values(values: list[str]) -> bool:
    return any(value for value in values)


def _collect_notes(values: list[str], start_index: int) -> str:
    return " | ".join(value for value in values[start_index:] if value)


def _find_label(values: list[str], aliases: dict[str, str]) -> tuple[int, str] | None:
    for index, value in enumerate(values):
        normalized = _normalize_label(value)
        if not normalized:
            continue
        canonical = aliases.get(normalized)
        if canonical:
            return index, canonical
    return None


def _value_after(values: list[str], start_index: int) -> str:
    if start_index + 1 < len(values):
        candidate = values[start_index + 1]
        if candidate:
            return candidate
    for candidate in values[start_index + 2 :]:
        if candidate:
            return candidate
    return ""


def _read_sheet_rows(sheet: openpyxl.worksheet.worksheet.Worksheet) -> list[tuple[int, list[str]]]:
    rows: list[tuple[int, list[str]]] = []
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=sheet.max_column, values_only=True),
        start=1,
    ):
        rows.append((row_number, [clean_cell(value) for value in row]))
    return rows


def _extract_sections(rows: Iterable[tuple[int, list[str]]]) -> tuple[dict[str, str], dict[str, list[tuple[int, list[str]]]]]:
    metadata: dict[str, str] = {}
    sections: dict[str, list[tuple[int, list[str]]]] = {section: [] for section in SECTION_LABELS}
    current_section: str | None = None

    for row_number, values in rows:
        if not _row_has_values(values):
            continue

        section_match = _find_label(values, SECTION_LABEL_ALIASES)
        if section_match:
            _, current_section = section_match
            continue

        metadata_match = _find_label(values, METADATA_FIELD_ALIASES)
        if metadata_match:
            index, canonical = metadata_match
            metadata[canonical] = _value_after(values, index)
            continue

        if current_section:
            sections[current_section].append((row_number, values))

    return metadata, sections


def _first_product(sections: dict[str, list[tuple[int, list[str]]]]) -> dict[str, str]:
    product_rows = sections.get("Products", []) or sections.get("Waste treatment", [])
    if not product_rows:
        return {"name": "", "quantity": "", "unit": ""}

    _, values = product_rows[0]
    return {
        "name": values[0] if len(values) > 0 else "",
        "quantity": values[1] if len(values) > 1 else "",
        "unit": values[2] if len(values) > 2 else "",
    }


def _map_section_row(section: str, values: list[str]) -> dict[str, str]:
    if section in {"Products", "Avoided products", "Waste treatment"}:
        amount = values[1] if len(values) > 1 else ""
        normalized = parse_quantity(amount)
        category = values[5] if len(values) > 5 else ""
        return {
            "dataset_utilizado": values[0] if len(values) > 0 else "",
            "compartimento": "",
            "cantidad_original": amount,
            "cantidad_normalizada": format_quantity(normalized),
            "unidad_medida": values[2] if len(values) > 2 else "",
            "incertidumbre": values[3] if len(values) > 3 else "",
            "comentario": category,
        }

    if section in {"Resources", "Emissions to air", "Emissions to water", "Emissions to soil"}:
        amount = values[2] if len(values) > 2 else ""
        normalized = parse_quantity(amount)
        return {
            "dataset_utilizado": values[0] if len(values) > 0 else "",
            "compartimento": values[1] if len(values) > 1 else "",
            "cantidad_original": amount,
            "cantidad_normalizada": format_quantity(normalized),
            "unidad_medida": values[3] if len(values) > 3 else "",
            "incertidumbre": values[4] if len(values) > 4 else "",
            "comentario": _collect_notes(values, 5),
        }

    amount = values[1] if len(values) > 1 else ""
    normalized = parse_quantity(amount)
    return {
        "dataset_utilizado": values[0] if len(values) > 0 else "",
        "compartimento": "",
        "cantidad_original": amount,
        "cantidad_normalizada": format_quantity(normalized),
        "unidad_medida": values[2] if len(values) > 2 else "",
        "incertidumbre": values[3] if len(values) > 3 else "",
        "comentario": _collect_notes(values, 4),
    }


def extract_workbook(workbook: UploadedWorkbook) -> list[dict[str, Any]]:
    path_info = split_uploaded_path(workbook.filename)
    source = openpyxl.load_workbook(io.BytesIO(workbook.content), read_only=True, data_only=True)
    output: list[dict[str, Any]] = []

    try:
        for sheet in source.worksheets:
            rows = _read_sheet_rows(sheet)
            metadata, sections = _extract_sections(rows)
            product = _first_product(sections)

            for section, section_rows in sections.items():
                for row_number, values in section_rows:
                    mapped = _map_section_row(section, values)
                    if not mapped["dataset_utilizado"]:
                        continue
                    output.append(
                        {
                            **path_info,
                            "process_identifier": metadata.get("Process identifier", ""),
                            "fecha_proceso": metadata.get("Date", ""),
                            "producto_principal": product["name"],
                            "cantidad_producto": product["quantity"],
                            "unidad_producto": product["unit"],
                            "clasificacion": SECTION_LABELS[section],
                            "seccion_origen": section,
                            "fila_origen": row_number,
                            **mapped,
                        }
                    )
    finally:
        source.close()

    return output


def extract_workbooks(workbooks: Iterable[UploadedWorkbook]) -> ExtractionResult:
    rows: list[dict[str, Any]] = []
    errors: list[dict[str, str]] = []
    total = 0
    processed = 0
    allowed_suffixes = (".xlsx", ".xlsm")

    for workbook in workbooks:
        total += 1
        filename = workbook.filename.lower()
        if not filename.endswith(allowed_suffixes) or Path(workbook.filename).name.startswith("~$"):
            continue
        try:
            extracted = extract_workbook(workbook)
            rows.extend(extracted)
            processed += 1
        except Exception as exc:
            errors.append({"archivo": workbook.filename, "error": str(exc)})

    rows.sort(
        key=lambda row: (
            row.get("carpeta", ""),
            row.get("subcarpeta", ""),
            row.get("archivo", ""),
            int(row.get("fila_origen", 0) or 0),
        )
    )
    summary = {
        "files_received": total,
        "files_processed": processed,
        "records": len(rows),
        "errors": len(errors),
    }
    return ExtractionResult(rows=rows, errors=errors, summary=summary)


def extract_paths(paths: Iterable[Path], base_dir: Path | None = None) -> ExtractionResult:
    workbooks: list[UploadedWorkbook] = []
    base = base_dir.resolve() if base_dir else None
    allowed_suffixes = {".xlsx", ".xlsm"}

    for path in paths:
        if path.name.startswith("~$") or path.suffix.lower() not in allowed_suffixes:
            continue
        resolved = path.resolve()
        if base:
            try:
                filename = str(resolved.relative_to(base))
            except ValueError:
                filename = path.name
        else:
            filename = path.name
        workbooks.append(UploadedWorkbook(filename=filename, content=resolved.read_bytes()))

    return extract_workbooks(workbooks)


def rows_to_csv(rows: Iterable[dict[str, Any]]) -> str:
    buffer = io.StringIO(newline="")
    writer = csv.DictWriter(buffer, fieldnames=OUTPUT_COLUMNS, delimiter=CSV_DELIMITER, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({column: row.get(column, "") for column in OUTPUT_COLUMNS})
    return buffer.getvalue()


def rows_to_xlsx(
    rows: Iterable[dict[str, Any]],
    errors: Iterable[dict[str, str]] | None = None,
    summary: dict[str, int] | None = None,
) -> bytes:
    row_list = list(rows)
    error_list = list(errors or [])
    summary_data = summary or {}
    workbook = openpyxl.Workbook()

    dataset_rows = [row for row in row_list if not _is_emission_or_waste(row)]
    emission_rows = [row for row in row_list if _is_emission_or_waste(row)]

    summary_sheet = workbook.active
    summary_sheet.title = "Resumen"
    _write_summary_sheet(summary_sheet, row_list, error_list, summary_data)

    dataset_sheet = workbook.create_sheet("Datasets")
    _write_table_sheet(dataset_sheet, EXCEL_COLUMNS, dataset_rows, "DatasetsTable")

    emissions_sheet = workbook.create_sheet("Emisiones")
    _write_table_sheet(emissions_sheet, EXCEL_COLUMNS, emission_rows, "EmisionesTable")

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _write_summary_sheet(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    rows: list[dict[str, Any]],
    errors: list[dict[str, str]],
    summary: dict[str, int],
) -> None:
    sheet.append(["Metrica", "Valor"])
    for label, value in [
        ("Archivos recibidos", summary.get("files_received", 0)),
        ("Archivos procesados", summary.get("files_processed", 0)),
        ("Registros extraidos", summary.get("records", len(rows))),
        ("Errores", summary.get("errors", len(errors))),
    ]:
        sheet.append([label, value])

    sheet.append([])
    sheet.append(["Registros por carpeta", "Cantidad"])
    for folder, count in sorted(Counter(row.get("carpeta", "") or "(sin carpeta)" for row in rows).items()):
        sheet.append([folder, count])

    sheet.append([])
    sheet.append(["Registros por clasificacion", "Cantidad"])
    for classification, count in sorted(Counter(row.get("clasificacion", "") for row in rows).items()):
        sheet.append([classification, count])

    _style_header_row(sheet, 1)
    for row_index in range(7, sheet.max_row + 1):
        if sheet.cell(row=row_index, column=1).value and sheet.cell(row=row_index, column=2).value == "Cantidad":
            _style_header_row(sheet, row_index)
    sheet.freeze_panes = "A2"
    _autosize_columns(sheet)


def _is_emission_or_waste(row: dict[str, Any]) -> bool:
    classification = str(row.get("clasificacion", "")).lower()
    source_section = str(row.get("seccion_origen", "")).lower()
    return "emisiones" in classification or source_section == "waste to treatment"


def _write_table_sheet(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    columns: list[tuple[str, str]],
    rows: list[dict[str, Any]],
    table_name: str,
) -> None:
    sheet.append([header for header, _ in columns])
    for row in rows:
        sheet.append([_excel_value(key, row.get(key, "")) for _, key in columns])

    _style_header_row(sheet, 1)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    if rows:
        reference = f"A1:{get_column_letter(len(columns))}{sheet.max_row}"
        table = Table(displayName=table_name, ref=reference)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    _autosize_columns(sheet)


def _excel_value(column: str, value: Any) -> Any:
    if column in {"cantidad_normalizada", "cantidad_producto"}:
        parsed = parse_quantity(value)
        return parsed if parsed is not None else value
    if column == "fila_origen":
        try:
            return int(value)
        except (TypeError, ValueError):
            return value
    return value


def _style_header_row(sheet: openpyxl.worksheet.worksheet.Worksheet, row_index: int) -> None:
    fill = PatternFill("solid", fgColor="184735")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[row_index]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize_columns(sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    wrap_columns = {"B", "C"}
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            text = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(text))
            if column_letter in wrap_columns:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 58)


def process_folder(input_dir: str, output_xlsx: str) -> ExtractionResult:
    root = Path(input_dir).expanduser().resolve()
    output = Path(output_xlsx).expanduser().resolve()
    allowed_suffixes = {".xlsx", ".xlsm"}
    paths = [
        path
        for path in root.rglob("*")
        if path.is_file() and path.resolve() != output and path.suffix.lower() in allowed_suffixes
    ]
    result = extract_paths(paths, base_dir=root)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_bytes(rows_to_xlsx(result.rows, result.errors, result.summary))
    return result
