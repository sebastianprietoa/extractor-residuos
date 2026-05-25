from __future__ import annotations

import tempfile
import unittest
from io import BytesIO
from pathlib import Path

import openpyxl

from app.simapro import extract_paths, parse_quantity, process_folder, rows_to_csv, rows_to_xlsx, split_uploaded_path


class QuantityParserTests(unittest.TestCase):
    def test_parses_numeric_expressions(self) -> None:
        self.assertEqual(parse_quantity("100+272"), 372)
        self.assertEqual(parse_quantity("6*1000"), 6000)
        self.assertEqual(parse_quantity("(1631+8692)*1000"), 10323000)
        self.assertEqual(parse_quantity("48000,00+6380,00"), 54380)

    def test_rejects_non_numeric_values(self) -> None:
        self.assertIsNone(parse_quantity("Indefinido"))
        self.assertIsNone(parse_quantity("1 kg"))
        self.assertIsNone(parse_quantity("__import__('os').system('dir')"))


class PathClassificationTests(unittest.TestCase):
    def test_strips_recursos_root_and_keeps_folder_context(self) -> None:
        info = split_uploaded_path("Recursos/Insumos/Combustibles/Diesel.XLSX")
        self.assertEqual(info["carpeta"], "Insumos")
        self.assertEqual(info["subcarpeta"], "Combustibles")
        self.assertEqual(info["archivo"], "Diesel.XLSX")


class ResourceExtractionTests(unittest.TestCase):
    def test_extracts_expected_records_from_sample_workbook(self) -> None:
        with tempfile.TemporaryDirectory(prefix="simapro_test_") as temp_dir:
            root = Path(temp_dir)
            workbook_path = root / "Recursos" / "Insumos" / "Combustibles" / "Diesel.XLSX"
            workbook_path.parent.mkdir(parents=True, exist_ok=True)
            _build_sample_workbook(workbook_path)

            result = extract_paths([workbook_path], base_dir=root)
            self.assertEqual(result.summary["files_processed"], 1)
            self.assertGreaterEqual(result.summary["records"], 2)
            self.assertFalse(result.errors)

            rows = result.rows
            diesel_rows = [row for row in rows if row["archivo"] == "Diesel.XLSX"]
            self.assertTrue(any(row["dataset_utilizado"] == "Diesel" for row in diesel_rows))
            self.assertTrue(any(row["unidad_medida"] == "l" for row in diesel_rows))

            csv_text = rows_to_csv(rows)
            self.assertIn("dataset_utilizado", csv_text)
            self.assertIn("Diesel.XLSX", csv_text)

            xlsx_bytes = rows_to_xlsx(rows, result.errors, result.summary)
            workbook = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=False, data_only=True)
            self.assertEqual(workbook.sheetnames, ["Resumen", "Datasets", "Emisiones"])
            self.assertEqual(workbook["Resumen"]["A1"].value, "Metrica")
            self.assertEqual(workbook["Datasets"]["C1"].value, "Dataset utilizado")
            self.assertGreater(workbook["Emisiones"].max_row, 1)
            workbook.close()

            output_path = root / "simapro_output.xlsx"
            processed = process_folder(str(root), str(output_path))
            self.assertTrue(output_path.exists())
            self.assertGreaterEqual(processed.summary["records"], 2)

    def test_extracts_records_from_second_sheet(self) -> None:
        with tempfile.TemporaryDirectory(prefix="simapro_test_sheet_") as temp_dir:
            root = Path(temp_dir)
            workbook_path = root / "Recursos" / "Insumos" / "Combustibles" / "Fuel.xlsm"
            workbook_path.parent.mkdir(parents=True, exist_ok=True)
            _build_workbook_with_data_in_second_sheet(workbook_path)

            result = extract_paths([workbook_path], base_dir=root)

            self.assertEqual(result.summary["files_processed"], 1)
            self.assertGreaterEqual(result.summary["records"], 1)
            self.assertFalse(result.errors)
            self.assertTrue(any(row["dataset_utilizado"] == "Diesel" for row in result.rows))

    def test_extracts_records_from_spanish_sections(self) -> None:
        with tempfile.TemporaryDirectory(prefix="simapro_test_spanish_") as temp_dir:
            root = Path(temp_dir)
            workbook_path = root / "Recursos" / "Insumos" / "Combustibles" / "Combustible.xlsm"
            workbook_path.parent.mkdir(parents=True, exist_ok=True)
            _build_spanish_workbook(workbook_path)

            result = extract_paths([workbook_path], base_dir=root)

            self.assertEqual(result.summary["files_processed"], 1)
            self.assertGreaterEqual(result.summary["records"], 2)
            self.assertFalse(result.errors)
            self.assertTrue(any(row["seccion_origen"] == "Products" for row in result.rows))


def _build_sample_workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    rows = [
        ["Process identifier", "PROC-001"],
        ["Date", "2025-05-19"],
        ["Products"],
        ["Diesel", "100+272", "l", "1", "", "Combustion"],
        ["Resources"],
        ["Fuel oil", "Fossil fuels", "6*1000", "kg", "2", "note"],
        ["Emissions to air"],
        ["CO2", "air", "3*2", "kg", "1", "note"],
    ]

    for row in rows:
        sheet.append(row)

    workbook.save(path)
    workbook.close()


def _build_workbook_with_data_in_second_sheet(path: Path) -> None:
    workbook = openpyxl.Workbook()
    first_sheet = workbook.active
    first_sheet.title = "Portada"
    first_sheet.append(["Solo", "portada"])

    second_sheet = workbook.create_sheet("Datos")
    rows = [
        ["Process identifier", "PROC-002"],
        ["Date", "2025-05-20"],
        ["Products"],
        ["Diesel", "100+272", "l", "1", "", "Combustion"],
        ["Emissions to air"],
        ["CO2", "air", "3*2", "kg", "1", "note"],
    ]

    for row in rows:
        second_sheet.append(row)

    workbook.save(path)
    workbook.close()


def _build_spanish_workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    rows = [
        ["Identificador del proceso", "PROC-ES-001"],
        ["Fecha", "2025-05-21"],
        ["Productos"],
        ["Diesel", "100+272", "l", "1", "", "Combustion"],
        ["Recursos"],
        ["Fuel oil", "Combustibles fosiles", "6*1000", "kg", "2", "note"],
        ["Emisiones al aire"],
        ["CO2", "air", "3*2", "kg", "1", "note"],
    ]

    for row in rows:
        sheet.append(row)

    workbook.save(path)
    workbook.close()


if __name__ == "__main__":
    unittest.main()
